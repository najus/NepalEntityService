"""
Migration Context for script execution.

This module provides the MigrationContext class, which is passed to migration
scripts and provides access to services, file reading helpers, and logging.
"""

import csv
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from nes2.database.entity_database import EntityDatabase
from nes2.services.publication.service import PublicationService
from nes2.services.scraping.service import ScrapingService
from nes2.services.search.service import SearchService

logger = logging.getLogger(__name__)


class MigrationContext:
    """
    Thin context object passed to migration scripts.

    Provides:
    - Direct access to publication, search, scraping, and database services
    - File reading helpers for CSV, JSON, and Excel files
    - Simple logging mechanism for migration progress

    The context is intentionally minimal to reduce maintenance burden and
    keep migration scripts simple. Services handle their own validation
    and error handling.

    Example:
        >>> async def migrate(context):
        ...     # Read data
        ...     data = context.read_csv("data.csv")
        ...
        ...     # Use services directly
        ...     for row in data:
        ...         entity = Entity(...)
        ...         await context.publication.create_entity(
        ...             entity_data=entity_data,
        ...             author_id="migration-001-test",
        ...             change_description="Import entity"
        ...         )
        ...
        ...     context.log(f"Imported {len(data)} entities")
    """

    def __init__(
        self,
        publication_service: PublicationService,
        search_service: SearchService,
        scraping_service: ScrapingService,
        db: EntityDatabase,
        migration_dir: Path,
    ):
        """
        Initialize the Migration Context.

        Args:
            publication_service: Service for creating/updating entities and relationships
            search_service: Service for searching and querying entities
            scraping_service: Service for data extraction and normalization
            db: Database for direct read access to entities
            migration_dir: Path to the migration folder containing the script
        """
        self.publication = publication_service
        self.search = search_service
        self.scraping = scraping_service
        self.db = db
        self._migration_dir = Path(migration_dir)
        self._logs: List[str] = []

        logger.debug(f"MigrationContext initialized for {self._migration_dir}")

    @property
    def migration_dir(self) -> Path:
        """
        Path to the migration folder.

        Use this to construct paths to data files within the migration folder.

        Returns:
            Path to the migration folder

        Example:
            >>> data_file = context.migration_dir / "data.csv"
        """
        return self._migration_dir

    @property
    def logs(self) -> List[str]:
        """
        Get all log messages generated during migration execution.

        Returns:
            List of log messages
        """
        return self._logs.copy()

    def log(self, message: str) -> None:
        """
        Log a message during migration execution.

        Messages are stored in the context for later retrieval and
        printed to the console during execution.

        Args:
            message: Log message to record

        Example:
            >>> context.log("Processing 100 entities...")
            >>> context.log(f"Imported {count} entities successfully")
        """
        self._logs.append(message)
        print(f"[Migration] {message}")
        logger.info(f"Migration log: {message}")

    def read_csv(self, filename: str) -> List[Dict[str, Any]]:
        """
        Read CSV file from migration folder.

        Reads a CSV file and returns it as a list of dictionaries, where
        each dictionary represents a row with column names as keys.

        Args:
            filename: Name of the CSV file (relative to migration folder)

        Returns:
            List of dictionaries, one per row

        Raises:
            FileNotFoundError: If the CSV file doesn't exist
            csv.Error: If the CSV file is malformed

        Example:
            >>> data = context.read_csv("entities.csv")
            >>> for row in data:
            ...     print(row["name"], row["type"])
        """
        file_path = self._migration_dir / filename

        if not file_path.exists():
            error_msg = f"CSV file not found: {filename}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        logger.debug(f"Reading CSV file: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                data = list(reader)

            logger.debug(f"Read {len(data)} rows from {filename}")
            return data

        except csv.Error as e:
            error_msg = f"Error reading CSV file {filename}: {e}"
            logger.error(error_msg)
            raise

        except Exception as e:
            error_msg = f"Unexpected error reading CSV file {filename}: {e}"
            logger.error(error_msg)
            raise

    def read_json(self, filename: str) -> Any:
        """
        Read JSON file from migration folder.

        Reads a JSON file and returns the parsed data structure
        (dict, list, or primitive value).

        Args:
            filename: Name of the JSON file (relative to migration folder)

        Returns:
            Parsed JSON data (dict, list, or primitive)

        Raises:
            FileNotFoundError: If the JSON file doesn't exist
            json.JSONDecodeError: If the JSON file is malformed

        Example:
            >>> data = context.read_json("config.json")
            >>> print(data["setting"])
        """
        file_path = self._migration_dir / filename

        if not file_path.exists():
            error_msg = f"JSON file not found: {filename}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        logger.debug(f"Reading JSON file: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            logger.debug(f"Successfully read JSON from {filename}")
            return data

        except json.JSONDecodeError as e:
            error_msg = f"Error parsing JSON file {filename}: {e}"
            logger.error(error_msg)
            raise

        except Exception as e:
            error_msg = f"Unexpected error reading JSON file {filename}: {e}"
            logger.error(error_msg)
            raise

    def read_excel(
        self, filename: str, sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Read Excel file from migration folder.

        Reads an Excel file and returns it as a list of dictionaries, where
        each dictionary represents a row with column names as keys.

        Requires openpyxl to be installed.

        Args:
            filename: Name of the Excel file (relative to migration folder)
            sheet_name: Name of the sheet to read (default: first sheet)

        Returns:
            List of dictionaries, one per row

        Raises:
            FileNotFoundError: If the Excel file doesn't exist
            ImportError: If openpyxl is not installed
            Exception: If the Excel file is malformed or sheet doesn't exist

        Example:
            >>> data = context.read_excel("data.xlsx", sheet_name="Sheet1")
            >>> for row in data:
            ...     print(row["name"], row["value"])
        """
        file_path = self._migration_dir / filename

        if not file_path.exists():
            error_msg = f"Excel file not found: {filename}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)

        logger.debug(f"Reading Excel file: {file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            error_msg = (
                "openpyxl is required to read Excel files. "
                "Install it with: pip install openpyxl"
            )
            logger.error(error_msg)
            raise ImportError(error_msg)

        try:
            # Load workbook
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

            # Get the sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Read data
            data = []
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                logger.debug(f"Excel file {filename} is empty")
                return data

            # First row is headers
            headers = rows[0]

            # Convert remaining rows to dictionaries
            for row in rows[1:]:
                row_dict = {}
                for i, header in enumerate(headers):
                    if header is not None:
                        row_dict[str(header)] = row[i] if i < len(row) else None
                data.append(row_dict)

            workbook.close()

            logger.debug(
                f"Read {len(data)} rows from {filename}"
                f"{f' (sheet: {sheet_name})' if sheet_name else ''}"
            )

            return data

        except Exception as e:
            error_msg = f"Error reading Excel file {filename}: {e}"
            logger.error(error_msg)
            raise
